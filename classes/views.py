
# پایتون استاندارد
import hashlib
import csv
import os
import re
from io import BytesIO
from zipfile import ZipFile

# Django
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.http import HttpResponse, JsonResponse
from django.conf import settings
from django.utils import timezone
from django.template.loader import render_to_string

# Third-party
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import qrcode

# Local apps
from .forms import SchoolClassForm
from .models import SchoolClass
from students.models import StudentProfile
from subjects.models import Subject
from grades.models import Grade, ExamType
from teachers.models import TeacherProfile


User = get_user_model()

def grade_entry(request):
    class_id = request.GET.get('class')
    exam_id = request.GET.get('exam_type')

    if not class_id or not exam_id:
        return HttpResponse("صنف یا امتحان انتخاب نشده است.", status=400)

    school_class = get_object_or_404(SchoolClass, id=class_id)
    exam = get_object_or_404(ExamType, id=exam_id)

    students = StudentProfile.objects.filter(classes=school_class).order_by('user__first_name')
    subjects = Subject.objects.filter(subjectgroup=school_class.subject_group)

    context = {
        "school_class": school_class,
        "exam": exam,
        "students": students,
        "subjects": subjects,
    }

    return render(request, "grades/grade/grade_entry.html", context)


def load_grade_attendance(request, pk):
    school_class = get_object_or_404(SchoolClass, id=pk)
    students = StudentProfile.objects.filter(classes=school_class)

    if request.method == "POST":
        exam_id = request.POST.get("exam_id")
        exam = get_object_or_404(ExamType, id=exam_id)

        context = {
            "school_class": school_class,
            "students": students,
            "exam_name": exam.name,
        }
        return render(request, "grades/grade/grade_entry.html", context)
        
    # اگر GET باشد چیزی برنگردان
    return HttpResponse(status=400)

@login_required
def create_class(request):
    user = request.user
    # ---------- فرم کلاس ----------
    if request.method == "POST":
        form = SchoolClassForm(request.POST, user=user)
        if form.is_valid():
            school_class = form.save(commit=False)
            school_class.created_by = user
            school_class.save()
            form.save_m2m()  # ذخیره دانش‌آموزان ManyToMany
            return redirect("class_list")
    else:
        form = SchoolClassForm(user=user)

    return render(request, "classes/class_form.html", {"form": form})


@login_required
def update_class(request, pk):
    school_class = get_object_or_404(SchoolClass, pk=pk)
    if request.method == "POST":
        form = SchoolClassForm(request.POST, instance=school_class, user=request.user)
        if form.is_valid():
            form.save()
            return redirect("class_list")
    else:
        form = SchoolClassForm(instance=school_class, user=request.user)
    return render(request, "classes/class_form.html", {"form": form})


@login_required
def class_list(request):
    # فقط کلاس‌هایی که کاربر ایجاد کرده
    classes = SchoolClass.objects.filter(created_by=request.user)

    # امتحانات کاربر جاری
    exams = ExamType.objects.filter(created_by=request.user)

    # اگر کاربر معلم است، امتحانات سوپرادمین‌ها را هم اضافه کن
    if hasattr(request.user, "teacherprofile"):
        super_admins = User.objects.filter(is_superuser=True)
        admin_exams = ExamType.objects.filter(created_by__in=super_admins)
        exams = (exams | admin_exams).distinct()

    return render(request, "classes/class_list.html", {
        "classes": classes,
        "exams": exams,
    })

    

@login_required
def delete_class(request, pk):
    school_class = get_object_or_404(SchoolClass, pk=pk, created_by=request.user)
    if request.method == "POST":
        school_class.delete()
        return redirect("class_list")
    return render(request, "classes/class_confirm_delete.html", {"class_instance": school_class})


@login_required
def export_class_zawabetA4(request, pk):
    # پیدا کردن کلاس
    school_class = get_object_or_404(SchoolClass, id=pk, created_by=request.user)
    
    students = school_class.students.all()

    # مسیر فایل قالب اکسل
    template_path = os.path.join(settings.BASE_DIR, 'static', 'excels', 'zawabetA4.xlsx')
    wb = load_workbook(template_path)
    ws = wb.active

    # قرار دادن نام صنف در سلول T5
    ws["X5"].value = school_class.name
    # شروع نوشتن داده‌ها از ردیف 10
    from openpyxl.utils import get_column_letter

    start_row = 10
    for i, student in enumerate(students, start=start_row):
        ws[f"C{i}"].value = f"{student.user.first_name} {student.user.last_name}"  # نام کامل
        ws[f"D{i}"].value = getattr(student, 'father_name', '')                     # نام پدر
        ws[f"E{i}"].value = getattr(student, 'grandfather_name', '')               # نام پدر کلان
        ws[f"F{i}"].value = getattr(student, 'student_number', '')                 # شماره اساس
        ws[f"F{i}"].value = getattr(student, 'student_number', '')                 # نام صنف

        


    # آماده‌سازی پاسخ برای دانلود
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_name = re.sub(r'[\\/*?:"<>|]', "_", school_class.name)
    filename = f"{safe_name}_students.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response




@login_required
def export_class_shoqaA4(request, pk):
    # پیدا کردن کلاس
    school_class = get_object_or_404(SchoolClass, id=pk, created_by=request.user)
    students = school_class.students.all()

    # مسیر فایل قالب اکسل
    template_path = os.path.join(settings.BASE_DIR, 'static', 'excels', 'shoqaA4.xlsx')
    wb = load_workbook(template_path)
    ws = wb.active

    # قرار دادن نام صنف در سلول T5
    ws["X5"].value = school_class.name  
    # شروع نوشتن داده‌ها از ردیف 10
    from openpyxl.utils import get_column_letter

    start_row = 10
    for i, student in enumerate(students, start=start_row):
        ws[f"C{i}"].value = f"{student.user.first_name} {student.user.last_name}"  # نام کامل
        ws[f"D{i}"].value = getattr(student, 'father_name', '')                     # نام پدر



    # آماده‌سازی پاسخ برای دانلود
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_name = re.sub(r'[\\/*?:"<>|]', "_", school_class.name)
    filename = f"{safe_name}_students.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

    
    




# --- Patch برای جلوگیری از خطای usedforsecurity در Python 3.8 ---
_real_md5 = hashlib.md5
def _patched_md5(*args, **kwargs):
    kwargs.pop('usedforsecurity', None)  # حذف آرگومان اضافی
    return _real_md5(*args, **kwargs)
hashlib.md5 = _patched_md5
# ---------------------------------------------------------------





@login_required
def class_print_shoqa(request, pk):
    school_class = get_object_or_404(SchoolClass, id=pk)
    students = StudentProfile.objects.filter(classes=school_class)

    if request.method == "POST":
        exam_id = request.POST.get("exam_id")
        exam = get_object_or_404(ExamType, id=exam_id)

        context = {
            "school_class": school_class,
            "students": students,
            "exam_name": exam.name,
        }
        return render(request, "classes/class_shoqa_a4.html", context)

    # اگر GET باشد چیزی برنگردان
    return HttpResponse(status=400)


@login_required
def class_print_zawabet(request, pk):
    school_class = get_object_or_404(SchoolClass, id=pk)
    students = StudentProfile.objects.filter(classes=school_class)

    if request.method == "POST":
        exam_id = request.POST.get("exam_id")
        exam = get_object_or_404(ExamType, id=exam_id)

        context = {
            "school_class": school_class,
            "students": students,
            "exam_name": exam.name,
        }
        return render(request, "classes/class_shoqa_zawabet_a4.html", context)



@login_required
def download_class_qr(request, class_id):
    # گرفتن شی صنف
    school_class = SchoolClass.objects.get(id=class_id)
    
    # گرفتن تمام شاگردان آن صنف
    students = school_class.students.all()  # فرض: relation many-to-many به StudentProfile

    # ایجاد حافظه برای فایل ZIP
    zip_buffer = BytesIO()
    
    with ZipFile(zip_buffer, "w") as zip_file:
        for student in students:
            # داده QR (مثلاً شماره شاگرد یا attendance_token)
            qr_data = str(student.attendance_token)
            qr_img = qrcode.make(qr_data)
            
            # ذخیره QR در حافظه موقت
            img_buffer = BytesIO()
            qr_img.save(img_buffer, format="PNG")
            img_buffer.seek(0)
            
            # اسم فایل داخل ZIP
            filename = f"student_{student.id}_{student.user.first_name}.png"
            
            # اضافه کردن به ZIP
            zip_file.writestr(filename, img_buffer.getvalue())

    zip_buffer.seek(0)
    
    # بازگرداندن ZIP به مرورگر
    response = HttpResponse(zip_buffer, content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="class_{school_class.id}_qrs.zip"'
    return response




@login_required
def class_report_cards(request, pk):
    school_class = get_object_or_404(SchoolClass, id=pk)
    students = StudentProfile.objects.filter(classes=school_class)

    # محاسبه نمره و تعیین درجه
    for student in students:
        grades = Grade.objects.filter(student=student)
        if grades.exists():
            total_score = sum([g.score for g in grades])
            avg_score = total_score / grades.count()
        else:
            avg_score = 0

        student.avg_score = avg_score  # برای استفاده در قالب

        # تعیین درجه بر اساس میانگین
        if avg_score >= 90:
            student.grade = 'الف'
        elif avg_score >= 75:
            student.grade = 'ب'
        elif avg_score >= 60:
            student.grade = 'ج'
        else:
            student.grade = 'د'

        
    context = {
        'school_class': school_class,
        'students': students,
        'today': timezone.now(),
    }

    return render(request, 'shared/student_report_card.html', context)





